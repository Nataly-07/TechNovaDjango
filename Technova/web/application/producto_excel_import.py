"""
Importación masiva de productos desde Excel (.xlsx).
Validación previa + transacción atómica (todo o nada).

Las columnas se resuelven por **nombre de encabezado** (fila 1), no por posición:
puedes insertar columnas extra sin desfasar el stock, siempre que el título reconocible
sea único (p. ej. «Stock Inicial»). Dos columnas que mapeen al mismo campo interno
provocan error explícito en lugar de sobrescribir en silencio.
"""

from __future__ import annotations

import logging
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from django.db import IntegrityError, transaction
from django.http import HttpResponse

from producto.domain.entities import ProductoEntidad
from producto.models import Producto, ProductoCatalogoExtra
from proveedor.models import Proveedor

logger = logging.getLogger(__name__)

MAX_FILAS = 5000
MAX_ARCHIVO_BYTES = 6 * 1024 * 1024
# Debe coincidir con Producto.imagen_url (URLField max_length) y validaciones del admin web.
MAX_URL_IMAGEN_CARACTERES = 500

_PRODUCTO_CATEGORIAS_ALTA_WEB = frozenset({"Celulares", "Portátiles"})
_PRODUCTO_MARCAS_ALTA_WEB = frozenset({"Apple", "Lenovo", "Motorola", "Xiaomi"})
_PRODUCTO_COLORES_ALTA_WEB = frozenset(
    {
        "Negro",
        "Blanco",
        "Gris",
        "Azul",
        "Rojo",
        "Dorado",
        "Plateado",
        "Verde",
        "Morado",
        "Rosa",
    }
)


def _categorias_alta_permitidas() -> set[str]:
    base = set(_PRODUCTO_CATEGORIAS_ALTA_WEB)
    extras = set(
        ProductoCatalogoExtra.objects.filter(tipo=ProductoCatalogoExtra.Tipo.CATEGORIA).values_list(
            "nombre", flat=True
        )
    )
    return base | extras


def _marcas_alta_permitidas() -> set[str]:
    base = set(_PRODUCTO_MARCAS_ALTA_WEB)
    extras = set(
        ProductoCatalogoExtra.objects.filter(tipo=ProductoCatalogoExtra.Tipo.MARCA).values_list(
            "nombre", flat=True
        )
    )
    return base | extras


def _normalizar_header(val: Any) -> str:
    """Clave comparable para emparejar con _HEADER_TO_CAMPO (sin depender del índice físico)."""
    s = ("" if val is None else str(val)).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    # Quitar todo espacio Unicode (NBSP, thin space, etc.), no solo U+0020
    return "".join(ch for ch in s if not ch.isspace())


# Cabeceras normalizadas (sin espacios, sin acentos) -> campo interno
_HEADER_TO_CAMPO = {
    "codigo": "codigo",
    "nombre": "nombre",
    "proveedor": "proveedor",
    "categoria": "categoria",
    "marca": "marca",
    "color": "color",
    "stockinicial": "stock_inicial",
    # Alias explícitos (no usar solo «stock»: podría confundirse con inventario operativo)
    "inventarioinicial": "stock_inicial",
    "cantidadinicial": "stock_inicial",
    "unidadesiniciales": "stock_inicial",
    "cantidadinicial": "stock_inicial",
    "unidadesiniciales": "stock_inicial",
    "costounitario": "costo_unitario",
    "margen(%)": "margen_pct",
    "margen%": "margen_pct",
    "precioventa": "precio_venta",
    "descripcion": "descripcion",
    "urlimagen": "url_imagen",
}

_CAMPOS_REQUERIDOS = frozenset(
    {
        "codigo",
        "nombre",
        "proveedor",
        "categoria",
        "marca",
        "color",
        "stock_inicial",
        "costo_unitario",
    }
)


@dataclass
class ErrorImportacionExcel(Exception):
    """Error de validación con fila Excel (1 = encabezado)."""

    mensaje: str
    fila: int | None = None

    def __str__(self) -> str:
        if self.fila is not None:
            return f"Fila {self.fila}: {self.mensaje}"
        return self.mensaje


def _resolver_catalogo(valor: str, permitidos: set[str], etiqueta: str) -> str:
    raw = " ".join((valor or "").strip().split())
    if not raw:
        raise ValueError(f"{etiqueta} es obligatorio.")
    for p in permitidos:
        if p.lower() == raw.lower():
            return p
    raise ValueError(f'{etiqueta} «{raw}» no coincide con un valor registrado en el sistema.')


def _resolver_proveedor(nombre: str) -> Proveedor:
    raw = " ".join((nombre or "").strip().split())
    if not raw:
        raise ValueError("Proveedor es obligatorio.")
    q = Proveedor.objects.filter(activo=True, nombre__iexact=raw).first()
    if q is None:
        raise ValueError(f'Proveedor «{raw}» no encontrado o inactivo.')
    return q


def _decimal_celda(val: Any, opcional: bool = False) -> Decimal | None:
    if val is None or (isinstance(val, str) and not str(val).strip()):
        if opcional:
            return None
        raise ValueError("Valor numérico obligatorio.")
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(",", ".")
    if not s:
        if opcional:
            return None
        raise ValueError("Valor numérico obligatorio.")
    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise ValueError("Número inválido.") from exc


def _int_stock(val: Any) -> int:
    """
    Convierte la celda de stock a int sin redondeos arbitrarios: debe ser un entero exacto.
    Excel/openpyxl suele devolver float; se usa Decimal(str) para evitar basura binaria (ej. 9.9999999991).
    """
    if val is None or val == "":
        raise ValueError("Stock inicial es obligatorio.")
    # bool es subclase de int en Python; no aceptar True/False como 1/0 por error de formato Excel
    if isinstance(val, bool):
        raise ValueError("Stock inicial inválido (valor booleano en la celda).")
    if isinstance(val, (datetime, date)):
        raise ValueError("Stock inicial inválido (la celda está formateada como fecha).")
    if isinstance(val, int):
        n = val
    elif isinstance(val, Decimal):
        if val != val.to_integral_value():
            raise ValueError("Stock debe ser un número entero (sin decimales).")
        n = int(val)
    elif isinstance(val, float):
        d = Decimal(str(val))
        if d != d.to_integral_value():
            raise ValueError("Stock debe ser un número entero (sin decimales).")
        n = int(d)
    else:
        s = str(val).strip()
        if not s:
            raise ValueError("Stock inicial es obligatorio.")
        try:
            d = Decimal(s.replace(",", ".").replace(" ", ""))
        except InvalidOperation as exc:
            raise ValueError("Stock inicial inválido.") from exc
        if d != d.to_integral_value():
            raise ValueError("Stock debe ser un número entero (sin decimales).")
        n = int(d)
    if n < 0:
        raise ValueError("Stock no puede ser negativo.")
    return n


def _parsear_filas_openpyxl(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ErrorImportacionExcel(
            "Falta la dependencia openpyxl. Instálala con: pip install openpyxl",
            fila=None,
        ) from exc

    if len(content) > MAX_ARCHIVO_BYTES:
        raise ErrorImportacionExcel("El archivo supera el tamaño máximo permitido (6 MB).", fila=None)

    bio = BytesIO(content)
    try:
        wb = load_workbook(bio, read_only=True, data_only=True)
    except Exception as exc:
        raise ErrorImportacionExcel(f"No se pudo leer el archivo Excel: {exc}", fila=None) from exc

    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            primera = next(rows_iter)
        except StopIteration:
            raise ErrorImportacionExcel("El archivo está vacío.", fila=1)

        col_map: dict[str, int] = {}
        for idx, cell in enumerate(primera):
            key = _normalizar_header(cell)
            campo = _HEADER_TO_CAMPO.get(key) if key else None
            if not campo:
                continue
            if campo in col_map:
                prev_idx = col_map[campo]
                raise ErrorImportacionExcel(
                    f"Hay dos columnas que mapean al mismo campo «{campo}» "
                    f"(columnas {prev_idx + 1} y {idx + 1}). Deja una sola o renombra el encabezado duplicado.",
                    fila=1,
                )
            col_map[campo] = idx

        faltan = _CAMPOS_REQUERIDOS - set(col_map.keys())
        if faltan:
            raise ErrorImportacionExcel(
                f"Faltan columnas obligatorias en el encabezado: {', '.join(sorted(faltan))}. "
                "Usa la plantilla «Descargar Plantilla».",
                fila=1,
            )

        filas: list[dict[str, Any]] = []
        excel_row = 1
        for tupla in rows_iter:
            excel_row += 1
            if excel_row > MAX_FILAS + 1:
                raise ErrorImportacionExcel(
                    f"Demasiadas filas (máximo {MAX_FILAS}). Divide el archivo.",
                    fila=excel_row,
                )

            def get(campo: str) -> Any:
                j = col_map.get(campo)
                if j is None:
                    return None
                if j >= len(tupla):
                    return None
                return tupla[j]

            codigo = get("codigo")
            celdas = [get(c) for c in col_map]
            fila_vacia = all(
                v is None or (isinstance(v, str) and not str(v).strip()) or v == ""
                for v in celdas
            )
            if fila_vacia:
                continue
            if codigo is None or str(codigo).strip() == "":
                raise ErrorImportacionExcel("El código es obligatorio.", fila=excel_row)

            filas.append(
                {
                    "_excel_row": excel_row,
                    "codigo": str(codigo).strip(),
                    "nombre": get("nombre"),
                    "proveedor": get("proveedor"),
                    "categoria": get("categoria"),
                    "marca": get("marca"),
                    "color": get("color"),
                    "stock_inicial": get("stock_inicial"),
                    "costo_unitario": get("costo_unitario"),
                    "margen_pct": get("margen_pct"),
                    "precio_venta": get("precio_venta"),
                    "descripcion": get("descripcion"),
                    "url_imagen": get("url_imagen"),
                }
            )
        return filas
    finally:
        wb.close()


def _validar_y_construir_entidades(
    filas: list[dict[str, Any]],
) -> list[ProductoEntidad]:
    if not filas:
        raise ErrorImportacionExcel("No hay filas de datos (solo encabezado o celdas vacías).", fila=2)

    categorias = _categorias_alta_permitidas()
    marcas = _marcas_alta_permitidas()
    vistos: set[str] = set()
    entidades: list[ProductoEntidad] = []

    for fila in filas:
        rnum = fila["_excel_row"]
        codigo = fila["codigo"][:50]
        if not codigo:
            raise ErrorImportacionExcel("El código es obligatorio.", fila=rnum)
        clave = codigo.lower()
        if clave in vistos:
            raise ErrorImportacionExcel(f'El código «{codigo}» está repetido en el archivo.', fila=rnum)
        vistos.add(clave)
        if Producto.objects.filter(codigo=codigo).exists():
            raise ErrorImportacionExcel("El código de producto ya existe.", fila=rnum)

        nombre = " ".join(str(fila["nombre"] or "").strip().split())
        if not nombre or len(nombre) > 120:
            raise ErrorImportacionExcel("El nombre es obligatorio (máx. 120 caracteres).", fila=rnum)

        try:
            prov = _resolver_proveedor(str(fila["proveedor"] or ""))
            cat = _resolver_catalogo(str(fila["categoria"] or ""), categorias, "Categoría")
            mar = _resolver_catalogo(str(fila["marca"] or ""), marcas, "Marca")
            col = _resolver_catalogo(str(fila["color"] or ""), _PRODUCTO_COLORES_ALTA_WEB, "Color")
        except ValueError as e:
            raise ErrorImportacionExcel(str(e), fila=rnum) from e

        try:
            # Misma cantidad que alta manual: va a stock y stock_inicial en el repositorio.
            stock = _int_stock(fila["stock_inicial"])
            costo = _decimal_celda(fila["costo_unitario"], opcional=False)
        except ValueError as e:
            raise ErrorImportacionExcel(str(e), fila=rnum) from e

        if costo is None or costo < 0:
            raise ErrorImportacionExcel("El costo unitario debe ser mayor o igual a 0.", fila=rnum)

        precio_raw = _decimal_celda(fila["precio_venta"], opcional=True)
        margen_raw = _decimal_celda(fila["margen_pct"], opcional=True)

        if precio_raw is not None and precio_raw > 0:
            precio = precio_raw.quantize(Decimal("0.01"))
        elif margen_raw is not None:
            precio = (costo * (Decimal("1") + margen_raw / Decimal("100"))).quantize(Decimal("0.01"))
        else:
            raise ErrorImportacionExcel(
                "Indica «Precio Venta» o «Margen (%)» para calcular el precio.",
                fila=rnum,
            )

        if precio <= 0:
            raise ErrorImportacionExcel("El precio de venta debe ser mayor que 0.", fila=rnum)
        if precio <= costo:
            raise ErrorImportacionExcel("El precio de venta debe ser mayor que el costo unitario.", fila=rnum)

        desc = " ".join(str(fila["descripcion"] or "").strip().split())
        url_img = str(fila["url_imagen"] or "").strip()
        if len(url_img) > MAX_URL_IMAGEN_CARACTERES:
            raise ErrorImportacionExcel(
                f"La URL de imagen es demasiado larga (máx. {MAX_URL_IMAGEN_CARACTERES} caracteres).",
                fila=rnum,
            )
        if url_img and not (url_img.startswith("http://") or url_img.startswith("https://")):
            raise ErrorImportacionExcel(
                "La URL de imagen debe comenzar con http:// o https://",
                fila=rnum,
            )

        entidades.append(
            ProductoEntidad(
                id=None,
                codigo=codigo,
                nombre=nombre,
                proveedor_id=prov.id,
                stock=stock,
                costo_unitario=costo,
                activo=True,
                imagen_url=url_img,
                categoria=cat,
                marca=mar,
                color=col,
                descripcion=desc,
                precio_venta=precio,
            )
        )
    return entidades


def importar_productos_desde_bytes(
    content: bytes,
    get_producto_service,
) -> int:
    """
    Valida todo el archivo, luego crea todos los productos en una sola transacción.
    Devuelve la cantidad creada.
    """
    filas = _parsear_filas_openpyxl(content)
    entidades = _validar_y_construir_entidades(filas)
    service = get_producto_service()

    with transaction.atomic():
        for ent in entidades:
            try:
                service.crear(ent)
            except IntegrityError as exc:
                raise ErrorImportacionExcel(
                    f"Error al guardar «{ent.codigo}»: posible código duplicado.",
                    fila=None,
                ) from exc

    return len(entidades)


def respuesta_plantilla_excel() -> HttpResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("openpyxl no está instalado.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    headers = [
        "Código",
        "Nombre",
        "Proveedor",
        "Categoría",
        "Marca",
        "Color",
        "Stock Inicial",
        "Costo Unitario",
        "Margen (%)",
        "Precio Venta",
        "Descripción",
        "URL Imagen",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)

    ejemplo = [
        "EJ-001",
        "Smartphone X",
        "Nombre del proveedor exacto",
        "Celulares",
        "Motorola",
        "Negro",
        10,
        500000,
        25,
        "",
        "Opcional: texto",
        "",
    ]
    for col, v in enumerate(ejemplo, 1):
        ws.cell(row=2, column=col, value=v)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="plantilla_productos_technova.xlsx"'
    return resp
